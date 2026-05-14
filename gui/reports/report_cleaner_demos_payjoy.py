# -*- coding: utf-8 -*-
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import polars as pl
import pandas as pd
import os, re

ID_COLUMNS_1 = ["financeorderid", "creditlineid"]
ID_COLUMNS_2 = ["adminname", "name"]

DEMOGRAPHIC_COLUMNS = [
    "phone","requiredsimphonenumber","phonenumber","phonenumber2",
    "requiredphonenumber","smsconfirmedphonenumber","email","tel contacto"
]

DELIMITERS = ["|", ",", "-", ";", "/", "\\", " | ", " - ", " , "]

DETAIL_HEADERS = {
    "source_file":"archivo_origen",
    "sheet":"hoja",
    "id":"id",
    "demo_column":"columna_demografico",
    "value":"valor",
    "type":"tipo",
}

SUMMARY_HEADERS = {
    "id":"ID",
    "assignments":"Asignaciones",
    "unique_mobile":"Celulares Únicos",
    "unique_landline":"Fijos Únicos",
    "unique_email":"Correos Únicos",
    "unique_invalid":"Inválidos Únicos",
    "total_valid":"Total Demográficos Válidos",
    "exclusive_demographics":"Demográficos Únicos Exclusivos",
    "shared_demographics":"Demográficos Compartidos con Otros IDs",
}

DEMO_VIEW_HEADERS = {
    "value":"Demográfico",
    "type":"Tipo",
    "total_assignments":"Total Asignaciones con este Demográfico",
    "total_occurrences":"Total Ocurrencias (sin quitar duplicados)",
    "unique_ids":"IDs Únicos que lo Tienen",
    "exclusivity":"Exclusividad",
}

TYPE_LABELS = {
    "mobile":"Celular",
    "landline":"Fijo",
    "email":"Correo",
    "invalid":"Inválido",
}

EXCEL_MAX_ROWS = 1048575

SPLIT_PATTERN = re.compile("|".join(re.escape(d) for d in sorted(DELIMITERS,key=len,reverse=True)))
NON_DIGIT = re.compile(r"\D")
NORMALIZE_COL = re.compile(r"[\s_\-]+")

def _split_multi(v):
    return [p.strip() for p in SPLIT_PATTERN.split(str(v)) if p.strip()]

def _classify(v):

    v = str(v).strip()

    if not v:
        return "", "invalid"

    if "@" in v:
        return v.lower(), "email"

    digits = NON_DIGIT.sub("", v)

    if digits.startswith("57") and len(digits) == 12:
        digits = digits[2:]

    if len(digits) == 10:

        if digits.startswith("3"):
            return digits, "mobile"

        if digits.startswith("6"):
            return digits, "landline"

    if len(digits) in (7,8):
        return digits, "landline"

    return v.lower(), "invalid"

def _clean_value(raw):

    if raw is None:
        return []

    raw = str(raw).strip()

    if not raw or raw == "nan":
        return []

    out = []

    for p in _split_multi(raw):

        cleaned, kind = _classify(p)

        if cleaned:
            out.append((cleaned, kind))

    return out

def _normalize_col(name):
    return NORMALIZE_COL.sub("", str(name).lower().strip())

def _find_id_col(cols, id_columns):

    norm = {_normalize_col(c): c for c in cols}

    for k in id_columns:
        if k in norm:
            return norm[k]

    return None

def _find_demo_cols(cols):

    norm = {_normalize_col(c): c for c in cols}

    return {
        k:norm[k]
        for k in DEMOGRAPHIC_COLUMNS
        if k in norm
    }

def process_excel(args):

    filepath, id_columns = args

    rows = []
    filename = filepath.stem

    try:
        xls = pd.ExcelFile(filepath, engine="calamine")

    except Exception as e:
        print(f"❌ Error {filepath.name}: {e}")
        return rows

    for sheet_name in xls.sheet_names:

        try:

            preview = pd.read_excel(
                filepath,
                sheet_name=sheet_name,
                nrows=0,
                engine="calamine"
            )

            preview.columns = [str(c) for c in preview.columns]

            id_col = _find_id_col(preview.columns, id_columns)
            demo_map = _find_demo_cols(preview.columns)

            if not id_col:
                continue

            usecols = [id_col] + list(set(demo_map.values()))

            df = pl.from_pandas(
                pd.read_excel(
                    filepath,
                    sheet_name=sheet_name,
                    usecols=usecols,
                    dtype=str,
                    engine="calamine"
                )
            )

        except Exception as e:
            print(f"❌ Error hoja {sheet_name} en {filepath.name}: {e}")
            continue

        cols = df.columns
        col_index = {c:i for i,c in enumerate(cols)}
        id_idx = col_index[id_col]

        demo_indexes = {
            c:col_index[c]
            for c in demo_map.values()
            if c in col_index
        }

        data = df.rows()

        print(f"✅ [{filepath.name} | {sheet_name}] {len(data):,} registros")

        for row in data:

            record_id = row[id_idx]

            if record_id is None:
                continue

            record_id = str(record_id).strip()

            if not record_id or record_id == "nan":
                continue

            for col_name, idx in demo_indexes.items():

                raw = row[idx]

                cleaned_list = _clean_value(raw)

                if not cleaned_list:
                    continue

                for cleaned, kind in cleaned_list:

                    rows.append((
                        filename,
                        sheet_name,
                        record_id,
                        col_name,
                        cleaned,
                        kind
                    ))

    return rows

def build_summary(df):

    valid = df.filter(pl.col("type") != "invalid")
    invalid = df.filter(pl.col("type") == "invalid")

    unique_files = df.group_by("id").agg(
        pl.col("source_file").n_unique().alias("assignments")
    )

    mobile = valid.filter(pl.col("type") == "mobile").group_by("id").agg(
        pl.col("value").n_unique().alias("unique_mobile")
    )

    landline = valid.filter(pl.col("type") == "landline").group_by("id").agg(
        pl.col("value").n_unique().alias("unique_landline")
    )

    email = valid.filter(pl.col("type") == "email").group_by("id").agg(
        pl.col("value").n_unique().alias("unique_email")
    )

    inv = invalid.group_by("id").agg(
        pl.col("value").n_unique().alias("unique_invalid")
    )

    summary = (
        unique_files
        .join(mobile,on="id",how="left")
        .join(landline,on="id",how="left")
        .join(email,on="id",how="left")
        .join(inv,on="id",how="left")
        .fill_null(0)
    )

    summary = summary.with_columns(
        (
            pl.col("unique_mobile") +
            pl.col("unique_landline") +
            pl.col("unique_email")
        ).alias("total_valid")
    )

    value_counts = valid.group_by("value").agg(
        pl.col("id").n_unique().alias("id_count")
    )

    exclusive = (
        valid
        .join(value_counts,on="value")
        .filter(pl.col("id_count") == 1)
        .group_by("id")
        .agg(pl.col("value").n_unique().alias("exclusive_demographics"))
    )

    summary = summary.join(exclusive,on="id",how="left").fill_null(0)

    summary = summary.with_columns(
        (
            pl.col("total_valid") -
            pl.col("exclusive_demographics")
        ).alias("shared_demographics")
    )

    return summary

def build_demo_view(df):

    valid = df.filter(pl.col("type") != "invalid")

    out = (
        valid
        .group_by(["value","type"])
        .agg([
            pl.col("source_file").n_unique().alias("total_assignments"),
            pl.len().alias("total_occurrences"),
            pl.col("id").n_unique().alias("unique_ids")
        ])
    )

    out = out.with_columns(
        pl.when(pl.col("unique_ids") > 1)
        .then(pl.lit("Repetido"))
        .otherwise(pl.lit("Único por ID"))
        .alias("exclusivity")
    )

    return out.sort(["type","unique_ids"], descending=[False,True])

def build_general_summary(df):

    valid = df.filter(pl.col("type") != "invalid")

    repeated_values = (
        valid
        .group_by("value")
        .agg(pl.col("id").n_unique().alias("ids"))
        .filter(pl.col("ids") > 1)
    )

    repeated_df = valid.join(repeated_values,on="value",how="inner")

    repeated_by_type = (
        repeated_df
        .unique("value")
        .group_by("type")
        .agg(pl.len().alias("count"))
    )

    repeated_dict = {
        r["type"]:r["count"]
        for r in repeated_by_type.to_dicts()
    }

    type_counts = (
        valid
        .unique(["id","value"])
        .group_by("type")
        .agg(pl.len().alias("count"))
    )

    type_dict = {
        r["type"]:r["count"]
        for r in type_counts.to_dicts()
    }

    ids_sharing = repeated_df.select(pl.col("id").n_unique()).item()
    total_unique_ids = df.select(pl.col("id").n_unique()).item()

    rows = [
        ("IDs únicos totales", total_unique_ids),
        ("Total de asignaciones analizadas", df.select(pl.col("source_file").n_unique()).item()),
        ("Total de demográficos únicos válidos", valid.select(pl.col("value").n_unique()).item()),
        ("Demográficos de correo repetidos en más de un ID", repeated_dict.get("email",0)),
        ("Demográficos de celular repetidos en más de un ID", repeated_dict.get("mobile",0)),
        ("Demográficos de fijo repetidos en más de un ID", repeated_dict.get("landline",0)),
        ("Total demográficos de correo", type_dict.get("email",0)),
        ("Total demográficos de celular", type_dict.get("mobile",0)),
        ("Total demográficos de fijo", type_dict.get("landline",0)),
        ("Total demográficos inválidos", df.filter(pl.col("type") == "invalid").select(pl.col("value").n_unique()).item()),
        ("IDs que comparten demográficos con otros IDs", ids_sharing),
        ("IDs con demográficos válidos solo propios (exclusivos)", total_unique_ids - ids_sharing),
    ]

    return pd.DataFrame(rows, columns=["Métrica","Valor"])

def _style_header(ws, cols, fill_hex="2563EB"):

    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)

    font = Font(bold=True,color="FFFFFF",name="Arial",size=10)

    align = Alignment(horizontal="center",vertical="center",wrap_text=True)

    for i, col in enumerate(cols, 1):

        cell = ws.cell(row=1,column=i,value=col)

        cell.fill = fill
        cell.font = font
        cell.alignment = align

    ws.row_dimensions[1].height = 25

def _freeze_top_row(ws):
    ws.freeze_panes = "A2"

def write_sheet(wb, sheet_name, df, headers=None, width=25, fill_hex="2563EB"):

    if isinstance(df, pl.DataFrame):
        df = df.to_pandas()

    if headers:
        df = df.rename(columns=headers)

    total_rows = len(df)
    chunks = max(1, (total_rows // EXCEL_MAX_ROWS) + 1)

    for chunk_idx in range(chunks):

        start = chunk_idx * EXCEL_MAX_ROWS
        end = start + EXCEL_MAX_ROWS

        chunk_df = df.iloc[start:end]

        current_name = (
            sheet_name
            if chunk_idx == 0
            else f"{sheet_name}_{chunk_idx + 1}"
        )

        ws = (
            wb.active
            if chunk_idx == 0 and wb.active.title == "Sheet"
            else wb.create_sheet(current_name)
        )

        ws.title = current_name

        cols = list(chunk_df.columns)

        _style_header(ws, cols, fill_hex)

        for r_idx, row in enumerate(chunk_df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx,column=c_idx,value=val)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = width

        _freeze_top_row(ws)

def write_output(detail_df, summary_df, demo_view_df, general_df, output_path):

    wb = Workbook()

    detail_pd = detail_df.to_pandas()
    detail_pd["type"] = detail_pd["type"].map(TYPE_LABELS)

    write_sheet(wb,"Detalle",detail_pd,DETAIL_HEADERS,22,"2563EB")
    write_sheet(wb,"Resumen por ID",summary_df,SUMMARY_HEADERS,26,"16A34A")

    demo_pd = demo_view_df.to_pandas()
    demo_pd["type"] = demo_pd["type"].map(TYPE_LABELS)

    write_sheet(wb,"Resumen por Demográfico",demo_pd,DEMO_VIEW_HEADERS,30,"7C3AED")
    write_sheet(wb,"Resumen General",general_df,None,52,"0F172A")

    wb.save(output_path)

def run_demographic_payjoy(input_folder, output_folder, id_columns, report_name):

    input_path = Path(input_folder)
    output_path = Path(output_folder)

    output_path.mkdir(parents=True, exist_ok=True)

    excel_files = [
        f for f in (
            list(input_path.rglob("*.xlsx")) +
            list(input_path.rglob("*.xls"))
        )
        if not f.name.startswith("~$")
    ]

    print(f"\n🚀 Procesando: {report_name}")
    print(f"📂 Archivos encontrados: {len(excel_files)}")

    all_rows = []

    tasks = [(fp, id_columns) for fp in excel_files]

    with ThreadPoolExecutor(max_workers=min(4, os.cpu_count())) as executor:

        results = executor.map(process_excel, tasks)

        for rows in results:
            all_rows.extend(rows)

    if not all_rows:
        print("❌ No se extrajo información.")
        return

    detail_df = pl.DataFrame(
        all_rows,
        schema=[
            "source_file",
            "sheet",
            "id",
            "demo_column",
            "value",
            "type"
        ],
        orient="row"
    ).unique([
        "source_file",
        "id",
        "demo_column",
        "value"
    ])

    summary_df = build_summary(detail_df)
    demo_view_df = build_demo_view(detail_df)
    general_df = build_general_summary(detail_df)

    out_file = (
        output_path /
        f"{report_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )

    write_output(
        detail_df,
        summary_df,
        demo_view_df,
        general_df,
        out_file
    )

    print(f"\n✅ Listo: {out_file}")
    print(f"📄 Filas detalle: {detail_df.height:,}")
    print(f"🆔 IDs únicos: {detail_df.select(pl.col('id').n_unique()).item():,}")

def cleaner_demo_payjoy(INPUT_FOLDER, OUTPUT_FOLDER):

    run_demographic_payjoy(
        INPUT_FOLDER,
        OUTPUT_FOLDER,
        ID_COLUMNS_1,
        "reporte_demográficos_cuenta"
    )

    run_demographic_payjoy(
        INPUT_FOLDER,
        OUTPUT_FOLDER,
        ID_COLUMNS_2,
        "reporte_demográficos_adminname"
    )