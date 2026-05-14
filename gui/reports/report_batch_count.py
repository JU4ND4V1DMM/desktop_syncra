import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import polars as pl
import os
from pathlib import Path
from datetime import datetime
import re

def normalize_cuenta(cuenta):
    if cuenta is None or pd.isna(cuenta):
        return None
    cuenta_str = str(cuenta).strip()
    
    if cuenta_str.endswith('.0-'):
        cuenta_str = cuenta_str[:-3]
    
    if cuenta_str.endswith('.0'):
        cuenta_str = cuenta_str[:-2]
    
    cuenta_str = re.sub(r'[-.]', '', cuenta_str)
    
    cuenta_str = re.sub(r'^0+', '', cuenta_str)
    
    if cuenta_str == '':
        return None
    return cuenta_str

def process_call_files(input_folder: str, output_folder: str):
    dataframes = []
    marcas_df = None
    marcas_filename = None
    total_records = 0
    
    files = sorted(Path(input_folder).glob("*"))
    
    # Buscar archivo de marcas
    for file in files:
        if file.suffix.lower() in ['.csv', '.parquet']:
            try:
                if file.suffix.lower() == '.csv':
                    df_temp = pl.read_csv(file, separator=';', infer_schema_length=1000, truncate_ragged_lines=True, ignore_errors=True)
                else:
                    df_temp = pl.read_parquet(file)
                
                if 'Marca_Asignada' in df_temp.columns and ('Cuenta_Next' in df_temp.columns or 'Numero_Referencia_de_Pago' in df_temp.columns):
                    print(f"\n📊 Archivo de marcas encontrado: {file.name}")
                    marcas_filename = file.name
                    
                    columnas_disponibles = []
                    if 'Cuenta_Next' in df_temp.columns:
                        columnas_disponibles.append('Cuenta_Next')
                    if 'Numero_Referencia_de_Pago' in df_temp.columns:
                        columnas_disponibles.append('Numero_Referencia_de_Pago')
                    
                    print(f"  Columnas disponibles para cruce: {columnas_disponibles}")
                    
                    # Preparar DataFrame de marcas con ambas columnas normalizadas
                    marcas_data = []
                    
                    if 'Cuenta_Next' in df_temp.columns:
                        temp_df = df_temp.select([
                            pl.col('Cuenta_Next').cast(pl.Utf8).alias('cuenta_original'),
                            pl.col('Marca_Asignada').cast(pl.Utf8)
                        ]).with_columns([
                            pl.col('cuenta_original').map_elements(normalize_cuenta, return_dtype=pl.Utf8).alias('cuenta_normalizada')
                        ]).drop_nulls().unique(subset=['cuenta_normalizada'])
                        
                        if temp_df.height > 0:
                            temp_df = temp_df.with_columns([
                                pl.lit('Cuenta_Next').alias('tipo_cruce')
                            ])
                            marcas_data.append(temp_df)
                    
                    if 'Numero_Referencia_de_Pago' in df_temp.columns:
                        temp_df = df_temp.select([
                            pl.col('Numero_Referencia_de_Pago').cast(pl.Utf8).alias('cuenta_original'),
                            pl.col('Marca_Asignada').cast(pl.Utf8)
                        ]).with_columns([
                            pl.col('cuenta_original').map_elements(normalize_cuenta, return_dtype=pl.Utf8).alias('cuenta_normalizada')
                        ]).drop_nulls().unique(subset=['cuenta_normalizada'])
                        
                        if temp_df.height > 0:
                            temp_df = temp_df.with_columns([
                                pl.lit('Numero_Referencia_de_Pago').alias('tipo_cruce')
                            ])
                            marcas_data.append(temp_df)
                    
                    if marcas_data:
                        marcas_df = pl.concat(marcas_data).unique(subset=['cuenta_normalizada'], keep='first')
                        print(f"  ✅ {marcas_df.height:,} registros de marcas cargados")
                    break
            except Exception as e:
                print(f"  Error al leer {file.name}: {e}")
                continue
    
    # Procesar archivos de llamadas
    for file in files:
        if file.suffix.lower() in ['.csv', '.parquet']:
            # Saltar el archivo de marcas
            if marcas_filename is not None and file.name == marcas_filename:
                continue
                
            print(f"Procesando: {file.name}")
            
            try:
                if file.suffix.lower() == '.csv':
                    try:
                        df = pl.read_csv(
                            file, 
                            separator=';',
                            infer_schema_length=10000,
                            truncate_ragged_lines=True,
                            ignore_errors=True
                        )
                    except:
                        df = pl.read_csv(
                            file,
                            infer_schema_length=10000,
                            truncate_ragged_lines=True,
                            ignore_errors=True
                        )
                else:
                    df = pl.read_parquet(file)
                
                required_cols = ['fechagestion', 'perfil', 'cuenta_promesa']
                if all(col in df.columns for col in required_cols):
                    records = df.height
                    total_records += records
                    
                    df_clean = df.select([
                        pl.col('fechagestion').cast(pl.Utf8),
                        pl.col('perfil').cast(pl.Utf8),
                        pl.col('cuenta_promesa').cast(pl.Utf8).alias('cuenta_promesa_original')
                    ]).with_columns([
                        pl.col('cuenta_promesa_original').map_elements(normalize_cuenta, return_dtype=pl.Utf8).alias('cuenta_normalizada')
                    ]).drop_nulls()
                    
                    if marcas_df is not None:
                        # Cruce con marcas
                        df_clean = df_clean.join(
                            marcas_df.select(['cuenta_normalizada', 'Marca_Asignada', 'tipo_cruce']),
                            on='cuenta_normalizada',
                            how='left'
                        )
                        df_clean = df_clean.drop('cuenta_normalizada')
                    else:
                        df_clean = df_clean.drop('cuenta_normalizada')
                        df_clean = df_clean.with_columns(pl.lit(None).cast(pl.Utf8).alias('Marca_Asignada'))
                        df_clean = df_clean.with_columns(pl.lit(None).cast(pl.Utf8).alias('tipo_cruce'))
                    
                    dataframes.append(df_clean)
                    print(f"  ✅ {records:,} registros válidos")
                else:
                    missing = [col for col in required_cols if col not in df.columns]
                    print(f"  ⚠️  Columnas faltantes en {file.name}: {missing}")
                    
            except Exception as e:
                print(f"  ❌ Error procesando {file.name}: {e}")
    
    if not dataframes:
        print("No se encontraron archivos válidos")
        return
    
    print(f"\nCombinando {len(dataframes)} archivos...")
    combined_df = pl.concat(dataframes)
    print(f"Total registros: {combined_df.height:,}")
    
    print("Procesando fechas...")
    df_with_dates = add_date_column(combined_df)
    
    if df_with_dates is None:
        print("Error al procesar fechas")
        return
    
    df_with_dates = df_with_dates.with_columns([
        pl.when(pl.col('perfil').str.contains("BLASTER CONTROLNEXT|IVR SAEM|IVR IPCOM"))
        .then(pl.lit("IVR"))
        .when(pl.col('perfil').str.contains("MENSAJERIA"))
        .then(pl.lit("SMS"))
        .when(pl.col('perfil').str.contains("CORREO"))
        .then(pl.lit("EMAIL"))
        .when(pl.col('perfil').str.contains("IAGEN"))
        .then(pl.lit("IAGEN"))
        .otherwise(pl.col('perfil'))
        .alias('herramienta')
    ])
    
    if 'Marca_Asignada' not in df_with_dates.columns:
        df_with_dates = df_with_dates.with_columns(pl.lit('SIN MARCA').cast(pl.Utf8).alias('Marca_Asignada'))
    else:
        df_with_dates = df_with_dates.with_columns([
            pl.col('Marca_Asignada').fill_null('SIN MARCA')
        ])
    
    if marcas_df is not None:
        print("Guardando detalle de registros SIN MARCA (no cruzaron con ninguna marca)...")
        sin_marca_df = df_with_dates.filter(pl.col('Marca_Asignada') == 'SIN MARCA')
        
        if sin_marca_df.height > 0:
            current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_output_path = os.path.join(output_folder, f"registros_sin_marca_{current_date}.csv")
            
            sin_marca_df.select([
                'fechagestion',
                'perfil',
                'cuenta_promesa_original',
                'herramienta'
            ]).write_csv(csv_output_path, separator=';')
            
            print(f"  ✅ {sin_marca_df.height:,} registros SIN MARCA guardados en: {csv_output_path}")
        else:
            print("  ✅ No hay registros SIN MARCA para guardar")
    
    print("Generando detalle general...")
    detalle_general = (df_with_dates
        .group_by(['fecha', 'perfil', 'herramienta', 'Marca_Asignada'])
        .agg(pl.len().alias('contador'))
        .sort(['fecha', 'herramienta', 'perfil', 'Marca_Asignada'])
    )
    
    print("Generando detalle sin duplicados...")
    detalle_unicos = (df_with_dates
        .unique(subset=['cuenta_promesa_original', 'fecha', 'herramienta', 'Marca_Asignada'])
        .group_by(['fecha', 'perfil', 'herramienta', 'Marca_Asignada'])
        .agg(pl.len().alias('contador'))
        .sort(['fecha', 'herramienta', 'perfil', 'Marca_Asignada'])
    )
    
    print("Generando resumen general por herramienta...")
    resumen_general = (detalle_general
        .group_by(['fecha', 'herramienta'])
        .agg(pl.sum('contador').alias('total'))
        .pivot(
            on='herramienta',
            index='fecha',
            values='total',
            aggregate_function='first'
        )
        .fill_null(0)
        .sort('fecha')
    )
    
    print("Generando resumen general por marca...")
    resumen_general_marcas = (detalle_general
        .group_by(['fecha', 'Marca_Asignada'])
        .agg(pl.sum('contador').alias('total'))
        .pivot(
            on='Marca_Asignada',
            index='fecha',
            values='total',
            aggregate_function='first'
        )
        .fill_null(0)
        .sort('fecha')
    )
    
    print("Generando resumen sin duplicados por herramienta...")
    resumen_unicos = (detalle_unicos
        .group_by(['fecha', 'herramienta'])
        .agg(pl.sum('contador').alias('total'))
        .pivot(
            on='herramienta',
            index='fecha',
            values='total',
            aggregate_function='first'
        )
        .fill_null(0)
        .sort('fecha')
    )
    
    print("Generando resumen sin duplicados por marca...")
    resumen_unicos_marcas = (detalle_unicos
        .group_by(['fecha', 'Marca_Asignada'])
        .agg(pl.sum('contador').alias('total'))
        .pivot(
            on='Marca_Asignada',
            index='fecha',
            values='total',
            aggregate_function='first'
        )
        .fill_null(0)
        .sort('fecha')
    )
    
    print("Generando detalle herramienta por marca...")
    detalle_general_herramienta_marca = (df_with_dates
        .group_by(['fecha', 'herramienta', 'Marca_Asignada'])
        .agg(pl.len().alias('contador'))
        .sort(['fecha', 'herramienta', 'Marca_Asignada'])
    )
    
    detalle_unicos_herramienta_marca = (df_with_dates
        .unique(subset=['cuenta_promesa_original', 'fecha', 'herramienta', 'Marca_Asignada'])
        .group_by(['fecha', 'herramienta', 'Marca_Asignada'])
        .agg(pl.len().alias('contador'))
        .sort(['fecha', 'herramienta', 'Marca_Asignada'])
    )
    
    herramientas_unicas = sorted(df_with_dates['herramienta'].unique().to_list())
    marcas_unicas = sorted(df_with_dates['Marca_Asignada'].unique().to_list())
    
    resumen_general_herramienta_vs_marca = []
    for fecha in sorted(df_with_dates['fecha'].unique().to_list()):
        row = {'fecha': fecha}
        df_fecha = detalle_general_herramienta_marca.filter(pl.col('fecha') == fecha)
        
        for herramienta in herramientas_unicas:
            for marca in marcas_unicas:
                valor = df_fecha.filter(
                    (pl.col('herramienta') == herramienta) & 
                    (pl.col('Marca_Asignada') == marca)
                ).select('contador').sum().item()
                col_name = f"{herramienta}_{marca}"
                row[col_name] = valor if valor else 0
        resumen_general_herramienta_vs_marca.append(row)
    
    resumen_general_herramienta_vs_marca = pl.DataFrame(resumen_general_herramienta_vs_marca).sort('fecha')
    
    resumen_unicos_herramienta_vs_marca = []
    for fecha in sorted(df_with_dates['fecha'].unique().to_list()):
        row = {'fecha': fecha}
        df_fecha = detalle_unicos_herramienta_marca.filter(pl.col('fecha') == fecha)
        
        for herramienta in herramientas_unicas:
            for marca in marcas_unicas:
                valor = df_fecha.filter(
                    (pl.col('herramienta') == herramienta) & 
                    (pl.col('Marca_Asignada') == marca)
                ).select('contador').sum().item()
                col_name = f"{herramienta}_{marca}"
                row[col_name] = valor if valor else 0
        resumen_unicos_herramienta_vs_marca.append(row)
    
    resumen_unicos_herramienta_vs_marca = pl.DataFrame(resumen_unicos_herramienta_vs_marca).sort('fecha')
    
    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"reporte_recuentos_batch_{current_date}.xlsx")
    
    print(f"Guardando: {output_path}")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        resumen_general.to_pandas().to_excel(writer, sheet_name='Resumen General', index=False)
        resumen_unicos.to_pandas().to_excel(writer, sheet_name='Resumen Unicos', index=False)
        detalle_general.to_pandas().to_excel(writer, sheet_name='Detalle General', index=False)
        detalle_unicos.to_pandas().to_excel(writer, sheet_name='Detalle Unicos', index=False)
        
        if marcas_df is not None:
            resumen_general_marcas.to_pandas().to_excel(writer, sheet_name='Resumen Gral x Marca', index=False)
            resumen_unicos_marcas.to_pandas().to_excel(writer, sheet_name='Resumen Unicos x Marca', index=False)
            resumen_general_herramienta_vs_marca.to_pandas().to_excel(writer, sheet_name='Detalle Gral Herram x Marca', index=False)
            resumen_unicos_herramienta_vs_marca.to_pandas().to_excel(writer, sheet_name='Detalle Unicos Herram x Marca', index=False)
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.row == 1:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin'),
                        right=openpyxl.styles.Side(style='thin'),
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    
                    if cell.row > 1 and cell.column > 1:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"\n✅ Proceso completado!")
    print(f"Total registros procesados: {combined_df.height:,}")
    print(f"Registros únicos: {df_with_dates.unique(subset=['cuenta_promesa_original', 'fecha', 'herramienta', 'Marca_Asignada']).height:,}")
    print(f"Rango de fechas: {df_with_dates['fecha'].min()} a {df_with_dates['fecha'].max()}")
    print(f"Herramientas: {resumen_general.columns[1:]}")
    if marcas_df is not None:
        print(f"Marcas encontradas: {sorted(df_with_dates['Marca_Asignada'].unique().to_list())}")
        print(f"Registros SIN MARCA (no cruzaron): {sin_marca_df.height if 'sin_marca_df' in locals() else 0:,}")
    
    return {
        'resumen_general': resumen_general,
        'resumen_unicos': resumen_unicos,
        'detalle_general': detalle_general,
        'detalle_unicos': detalle_unicos,
        'resumen_general_marcas': resumen_general_marcas if marcas_df is not None else None,
        'resumen_unicos_marcas': resumen_unicos_marcas if marcas_df is not None else None
    }

def add_date_column(df):
    date_formats = [
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
    ]
    
    for date_format in date_formats:
        try:
            df_with_dates = df.with_columns([
                pl.col('fechagestion')
                .str.to_datetime(format=date_format, strict=False)
                .dt.date()
                .alias('fecha')
            ]).drop_nulls(['fecha', 'perfil', 'cuenta_promesa_original'])
            
            if df_with_dates.height > 0:
                return df_with_dates
        except:
            continue
    
    try:
        df_with_dates = df.with_columns([
            pl.col('fechagestion')
            .str.slice(0, 10)
            .alias('fecha')
        ]).drop_nulls(['fecha', 'perfil', 'cuenta_promesa_original'])
        return df_with_dates
    except:
        return None