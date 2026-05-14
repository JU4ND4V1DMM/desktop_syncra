import os
import csv
from openpyxl import load_workbook
from datetime import datetime
import chardet
import pandas as pd

def details_files(folder_path, output_folder):
    # Crear ruta de salida válida
    folder_name = os.path.basename(folder_path)
    output_file = os.path.join(output_folder, f"Validación Archivos {folder_name} {datetime.now().strftime('%Y-%m-%d')}.csv")
    
    print(f"🔍 Iniciando procesamiento de: {folder_path}")
    print(f"📁 Carpeta: {folder_name}")
    print(f"💾 Archivo de salida: {output_file}")
    print("─" * 50)
    
    # Abrir archivo de salida
    with open(output_file, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['📄 Archivo', '📑 Hoja/Sección', '🏷️ Títulos/Columnas', '🔢 Total Filas', 'ℹ️ Observaciones'])
        
        total_files = 0
        processed_files = 0
        
        # Contar archivos primero
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.xlsx', '.xls', '.csv', '.txt')):
                total_files += 1
        
        print(f"📊 Total de archivos a procesar: {total_files}")
        
        # Procesar archivos
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            
            # Procesar archivos Excel
            if filename.lower().endswith(('.xlsx', '.xls')):
                print(f"\n📊 Procesando Excel: {filename}")
                processed_files += 1
                
                try:
                    workbook = load_workbook(file_path, read_only=True, data_only=True)
                    sheet_names = workbook.sheetnames
                    
                    for sheet_name in sheet_names:
                        print(f"   📑 Hoja: {sheet_name}")
                        sheet = workbook[sheet_name]
                        
                        # Obtener primera fila para títulos
                        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                        
                        # Contar filas totales en la hoja
                        try:
                            # Método para contar filas con datos
                            total_rows = 0
                            for row in sheet.iter_rows(values_only=True):
                                if any(cell is not None for cell in row):
                                    total_rows += 1
                            print(f"      📊 Filas totales: {total_rows}")
                        except:
                            # Fallback: usar max_row
                            total_rows = sheet.max_row
                            print(f"      📊 Filas totales (estimado): {total_rows}")
                        
                        if first_row:
                            titles = [str(cell) if cell is not None else '' for cell in first_row]
                            titles_text = ', '.join(titles[:50])
                            if len(titles) > 50:
                                titles_text += f" ... (+{len(titles)-50} más)"
                            writer.writerow([filename, sheet_name, titles_text, total_rows, '✅ Excel procesado'])
                        else:
                            writer.writerow([filename, sheet_name, "Sin títulos detectados", total_rows, '⚠️ Hoja vacía'])
                            
                except Exception as e:
                    print(f"   ❌ Error procesando Excel: {str(e)}")
                    writer.writerow([filename, 'ERROR', f'Error: {str(e)}', 0, '❌ Error en archivo'])
            
            # Procesar archivos CSV/TXT
            elif filename.lower().endswith(('.csv', '.txt')):
                print(f"\n📄 Procesando CSV/TXT: {filename}")
                processed_files += 1
                
                try:
                    # Detectar codificación
                    with open(file_path, 'rb') as f:
                        raw_data = f.read(10000)  # Leer primeros 10KB para detectar
                        result = chardet.detect(raw_data)
                        encoding = result['encoding']
                    
                    print(f"   🔤 Codificación detectada: {encoding}")
                    
                    # Intentar diferentes delimitadores
                    delimiters = [',', ';', '\t', '|', '~']
                    delimiter_found = None
                    sample_lines = []
                    
                    # Leer primeras líneas para análisis
                    with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                        for _ in range(5):
                            line = f.readline()
                            if line:
                                sample_lines.append(line)
                    
                    # Detectar mejor delimitador
                    best_delimiter = ','
                    max_columns = 0
                    
                    for delim in delimiters:
                        if sample_lines:
                            columns = sample_lines[0].count(delim)
                            if columns > max_columns:
                                max_columns = columns
                                best_delimiter = delim
                    
                    delimiter_found = best_delimiter
                    print(f"   🎯 Delimitador detectado: {repr(delimiter_found)}")
                    
                    # Leer CSV con delimitador detectado
                    try:
                        # Leer para contar filas y obtener títulos
                        df = pd.read_csv(file_path, 
                                        delimiter=delimiter_found,
                                        encoding=encoding,
                                        engine='python')
                        
                        total_rows = len(df)
                        print(f"      📊 Filas totales: {total_rows}")
                        
                        titles = list(df.columns)
                        titles_text = ', '.join([str(t)[:50] for t in titles[:8]])
                        if len(titles) > 8:
                            titles_text += f" ... (+{len(titles)-8} más)"
                        
                        writer.writerow([filename, 'CSV/TXT', titles_text, total_rows,
                                       f'✅ Codificación: {encoding}, Delimitador: {repr(delimiter_found)}'])
                        
                    except Exception as e:
                        print(f"   ⚠️ Probando métodos alternativos...")
                        # Intentar con csv.reader para contar filas manualmente
                        with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                            reader = csv.reader(f, delimiter=delimiter_found)
                            rows = list(reader)
                            total_rows = len(rows)
                            
                            try:
                                first_row = rows[0] if rows else []
                                titles_text = ', '.join([str(t)[:50] for t in first_row[:8]])
                                if len(first_row) > 8:
                                    titles_text += f" ... (+{len(first_row)-8} más)"
                                
                                writer.writerow([filename, 'CSV/TXT', titles_text, total_rows,
                                               f'✅ Codificación: {encoding}, Delimitador: {repr(delimiter_found)}'])
                            except IndexError:
                                writer.writerow([filename, 'CSV/TXT', 'Archivo vacío', 0, 
                                               '⚠️ Sin contenido'])
                            
                except Exception as e:
                    print(f"   ❌ Error procesando CSV: {str(e)}")
                    writer.writerow([filename, 'CSV/TXT', f'Error: {str(e)}', 0, '❌ Error en archivo'])
            
            else:
                # Ignorar otros tipos de archivos
                continue
        
        print("\n" + "═" * 50)
        print(f"✅ Procesamiento completado!")
        print(f"📈 Archivos procesados: {processed_files}/{total_files}")
        print(f"💾 Resultados guardados en: {output_file}")
    
    return output_file